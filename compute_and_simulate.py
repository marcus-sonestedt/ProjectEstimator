import sys
import openpyxl
import numpy as np
import plotly.graph_objects as go
from scipy.stats import gamma
from scipy.optimize import minimize

# Parameters
days = 100
percentages = { 'O': 0.30, 'L':0.50, 'P':0.95}
num_simulations = 250

def bucketify(s:str):
    try:
        if isinstance(s, int):
            return {s:1}

        votes = [int(v) for v in s.strip().split(' ')]
        histogram, bin_edges = np.histogram(votes, bins='auto')
        buckets = {}
        for i, count in enumerate(histogram):
            bucket_key = f"{bin_edges[i]:.1f}-{bin_edges[i+1]:.1f}"
            buckets[bucket_key] = int(count)
        
        return buckets
    except Exception as e:
        print(f"Error parsing votes '{s}': {e}")
        raise

class Task:
    def __init__(self, name, o_est, l_est, p_est):
        self.name = name
        self.estimations = {'O':bucketify(o_est),'L':bucketify(l_est),'P':bucketify(p_est)}

        # Extract the median values from each estimate
        self.o_value = int(o_est.split()[0]) if isinstance(o_est, str) else int(o_est)
        self.l_value = int(l_est.split()[0]) if isinstance(l_est, str) else int(l_est)
        self.p_value = int(p_est.split()[0]) if isinstance(p_est, str) else int(p_est)

        # Simple check 
        if self.o_value > self.l_value or self.l_value > self.p_value:
            print(f"Warning for task '{self.name}': OLP values should be non-decreasing (O={self.o_value}, L={self.l_value}, P={self.p_value})")        
        
        # Fit gamma distribution parameters
        self.compute_gamma_params()
    
    def compute_gamma_params(self):
        """Compute gamma distribution parameters based on full distribution of O, L, P values"""
        # Extract all vote values and their weights from the buckets
        o_values = []
        l_values = []
        p_values = []
        
        # Process the O, L, P buckets to create weighted arrays of values
        for est_type, buckets in [('O', self.estimations['O']), ('L', self.estimations['L']), ('P', self.estimations['P'])]:
            values_array = []
            
            # For each bucket, parse the range and extract values
            for bucket_key, count in buckets.items():
                # Handle the case where the bucket_key is a single integer
                if isinstance(bucket_key, int):
                    bucket_value = bucket_key
                else:
                    # Parse range like "10.0-15.0" to get midpoint
                    try:
                        low, high = bucket_key.split('-')
                        bucket_value = (float(low) + float(high)) / 2
                    except ValueError:
                        # If parsing fails, use first number in the string
                        bucket_value = float(bucket_key.split('-')[0])
                
                # Add values to the appropriate array with weights
                if est_type == 'O':
                    o_values.extend([bucket_value] * count)
                elif est_type == 'L':
                    l_values.extend([bucket_value] * count)
                else:  # 'P'
                    p_values.extend([bucket_value] * count)
        
        # Calculate percentiles from actual distributions
        if o_values and l_values and p_values:
            # Sort the arrays
            o_values.sort()
            l_values.sort()
            p_values.sort()
            
            # Get representative values at the specified percentiles
            o_percentile_idx = min(int(len(o_values) * percentages['O']), len(o_values) - 1)
            l_percentile_idx = min(int(len(l_values) * percentages['L']), len(l_values) - 1) 
            p_percentile_idx = min(int(len(p_values) * percentages['P']), len(p_values) - 1)
            
            # Use these as target values for gamma distribution fitting
            o_target = o_values[o_percentile_idx]
            l_target = l_values[l_percentile_idx]
            p_target = p_values[p_percentile_idx]
            
            # Store these for later use (they're more accurate than single values)
            self.o_value = o_target
            self.l_value = l_target
            self.p_value = p_target
        else:
            # Fall back to the first values if buckets are empty
            print(f"Warning: Insufficient vote data for task '{self.name}', using first values")
            # o_value, l_value, and p_value were already set in __init__
        
        # Calculate mean and variance using the full distribution
        all_values = o_values + l_values + p_values
        if all_values:
            weighted_mean = np.mean(all_values)
            weighted_variance = np.var(all_values)
        else:
            # Fall back to PERT formula if no values are available
            weighted_mean = (self.o_value + 4*self.l_value + self.p_value) / 6
            weighted_variance = ((self.p_value - self.o_value) / 6)**2
        
        # Rest of the method remains similar, but use weighted_mean and weighted_variance
        if weighted_variance < 0.001:
            # If variance is essentially zero, use a very narrow distribution
            self.gamma_alpha = 100.0  # High alpha for narrow distribution
            self.gamma_beta = 100.0 / weighted_mean if weighted_mean > 0 else 100.0
            return
        
        # Calculate initial guess parameters for gamma distribution
        # alpha = shape, beta = rate (1/scale)
        alpha_guess = (weighted_mean**2) / weighted_variance
        beta_guess = weighted_mean / weighted_variance
        
        # Define the objective function to minimize
        def objective(params):
            alpha, beta = params
            # Calculate squared errors between the desired percentiles and the actual ones
            o_error = (gamma.ppf(percentages['O'], alpha, scale=1/beta) - self.o_value)**2
            l_error = (gamma.ppf(percentages['L'], alpha, scale=1/beta) - self.l_value)**2
            p_error = (gamma.ppf(percentages['P'], alpha, scale=1/beta) - self.p_value)**2
            return o_error + l_error + p_error
        
        # Ensure initial guesses are positive and reasonable
        initial_guess = [max(0.1, alpha_guess), max(0.1, beta_guess)]
        
        # Minimize the objective function to find best alpha, beta
        try:
            result = minimize(objective, initial_guess, method='Nelder-Mead', 
                              options={'maxiter': 1000})
            
            if result.success:
                self.gamma_alpha, self.gamma_beta = result.x
                # Ensure parameters are positive
                self.gamma_alpha = max(0.1, self.gamma_alpha)
                self.gamma_beta = max(0.1, self.gamma_beta)
            else:
                # If optimization doesn't converge, fall back to initial guess
                self.gamma_alpha = max(0.1, alpha_guess)
                self.gamma_beta = max(0.1, beta_guess)
                print(f"Warning: Optimization failed for task '{self.name}', using fallback parameters")
                
        except Exception as e:
            # Fallback if optimization fails completely
            self.gamma_alpha = max(0.1, (weighted_mean**2) / weighted_variance if weighted_variance > 0.001 else 100.0)
            self.gamma_beta = max(0.1, weighted_mean / weighted_variance if weighted_variance > 0.001 else 100.0 / weighted_mean if weighted_mean > 0 else 100.0)
            print(f"Warning: Error fitting distribution for task '{self.name}': {e}")

    def sample_duration(self, size=1):
        """Sample task duration from gamma distribution"""
        return gamma.rvs(self.gamma_alpha, scale=1/self.gamma_beta, size=size)

    def __str__(self):
        return f'{self.name} {self.o_value} {self.l_value} {self.p_value}'
    
class TaskCompletionEstimator:
    def __init__(self):
        self.tasks = []
        
    def read(self, xls_file:str):
        print(f"Reading Excel file '{xls_file}'...")

        # Open the Excel workbook using openpyxl with read-only and data-only modes for shared access
        try:
            workbook = openpyxl.load_workbook(filename=xls_file, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]  # get the first worksheet

            # Read all rows and create tasks
            # In read-only mode, we need to convert to list to filter properly
            rows = list(sheet.iter_rows(values_only=True))
            self.rows = [row for row in rows if row[2] is not None][4:]
            self.tasks = [Task(name=row[1], o_est=row[2], l_est=row[3], p_est=row[4]) for row in self.rows]
        
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            raise

    def simulate(self):        
        self.completion_times = []
        
        print(f"Simulating project completion times ({num_simulations} iterations)...")
        # Run multiple simulations
        for _ in range(num_simulations):
            # Sample duration for each task
            task_durations = [task.sample_duration()[0] for task in self.tasks]
            
            # Project completion is sum of all task durations
            # (assuming tasks are sequential)
            total_duration = sum(task_durations)
            self.completion_times.append(total_duration)
        
        # Sort completion times for percentile calculation
        self.completion_times.sort()
        
        # Create histogram with bins at 0.5 day intervals
        max_completion_time = max(self.completion_times) + 1
        bin_edges = np.arange(0, max_completion_time + 0.5, 0.5)
        hist, bin_edges = np.histogram(self.completion_times, bins=bin_edges)
        self.cum_prob = np.cumsum(hist) / num_simulations
        self.bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2

    def plot(self):
        """Plot the cumulative probability of project completion using Plotly"""
        # Create figure
        fig = go.Figure()
        
        # Add cumulative probability curve with reduced "wiggliness"
        fig.add_trace(go.Scatter(
            x=self.bin_centers[:len(self.cum_prob)],
            y=self.cum_prob,
            mode='lines',
            name='Cumulative Probability',
            line=dict(
                color='royalblue', 
                width=2,
                shape='spline',      # Apply spline interpolation for smooth curve
                smoothing=1.0        # Reduced smoothing intensity to avoid too much wiggle
            ),
        ))
        
        # Add data points at 0.5 day intervals (all data points)
        fig.add_trace(go.Scatter(
            x=self.bin_centers[:len(self.cum_prob)],
            y=self.cum_prob,
            mode='markers',
            name='Data Points',
            marker=dict(color='royalblue', size=3),  # Smaller markers for better visibility
            showlegend=False
        ))
        
        # Add reference lines for key percentiles
        percentile_days = {}
        annotations = []
        
        for pct_name, pct_value in percentages.items():
            day_index = np.searchsorted(self.cum_prob, pct_value)
            if day_index < len(self.cum_prob):
                actual_day = self.bin_centers[day_index]
                percentile_days[pct_name] = actual_day
                
                # Add vertical line
                fig.add_vline(x=actual_day, line_dash="dash", line_color="firebrick", opacity=0.5)
                
                # Add horizontal line
                fig.add_hline(y=pct_value, line_dash="dash", line_color="firebrick", opacity=0.5)
                
                # Add point marker
                fig.add_trace(go.Scatter(
                    x=[actual_day],
                    y=[pct_value],
                    mode='markers',
                    marker=dict(color='firebrick', size=10),
                    name=f"{pct_name} ({pct_value*100:.0f}%)"
                ))
                
                # Add annotation
                annotations.append(dict(
                    x=actual_day + 1,
                    y=pct_value - 0.05,
                    text=f"{pct_name}: {actual_day:.1f} days",
                    showarrow=False,
                    font=dict(color="firebrick")
                ))
        
        fig.update_layout(
            title="Project Completion Probability",
            xaxis_title="Days",
            yaxis_title="Cumulative Probability of Completion",
            yaxis=dict(range=[0, 1]),
            annotations=annotations,
            template="plotly_white",
            hovermode="x unified"
        )
        
        fig.show()
        
        # Print the results
        print("Project completion estimates:")
        for pct_name, days in percentile_days.items():
            print(f"  {pct_name} ({percentages[pct_name]*100}%): {days:.1f} days")

if __name__ == '__main__':
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = r'Estimation 2025-02-25.xlsx'

    probe = TaskCompletionEstimator()
    probe.read(file_path)
    probe.simulate()
    probe.plot()

